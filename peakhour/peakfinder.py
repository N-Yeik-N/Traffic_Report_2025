from openpyxl import load_workbook
import warnings
from dataclasses import dataclass
import re
from datetime import datetime
from collections import Counter

warnings.filterwarnings("ignore")

""" 
Madrugada:  00:00 - 06:00
Mañana:     06:00 - 12:00 
Tarde:      12:00 - 17:00
Noche:      17:00 - 00:00
"""

@dataclass
class Hours:
    name: str
    fecha: datetime
    codigo: str
    id_morning: int
    id_evening: int
    id_night: int
    vol_morning: int
    vol_evening: int
    vol_night: int

def peakhour_finder(path) -> Hours:
    """ Devuelve las horas puntas para los 3 periodos junto a sus volumenes. """
    pattern1 = r'([A-Z]+-[0-9]+)'
    pattern2 = r'([A-Z]+[0-9]+)'

    match1 = re.search(pattern1, path)
    match2 = re.search(pattern2, path)

    if match1: code = match1.group(1)
    elif match2: code = match2.group(1)
    else: print("ERROR - Excel sin código de tipo AA-99 o AA99:\n",path)

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb['Inicio']
    intersection_name = ws['G5'].value
    date = ws['G6'].value

    ws = wb['N']
    hour_sum_slice = slice("HR16", "HR111")
    SUM_HOUR = [row[0].value for row in ws[hour_sum_slice]]
    wb.close()

    peakhour_morning = max(SUM_HOUR[6*4:12*4])
    index_morning = SUM_HOUR.index(peakhour_morning)
    volume_morning = SUM_HOUR[index_morning]
    
    peakhour_evening = max(SUM_HOUR[12*4:17*4])
    index_evening = SUM_HOUR.index(peakhour_evening)
    volume_evening = SUM_HOUR[index_evening]

    peakhour_night = max(SUM_HOUR[17*4:24*4])
    index_night = SUM_HOUR.index(peakhour_night)
    volume_night = SUM_HOUR[index_night]

    excel_hours_info = Hours(
        name = intersection_name,
        fecha = date,
        codigo = code,
        id_morning = index_morning,
        id_evening = index_evening,
        id_night = index_night,
        vol_morning = volume_morning,
        vol_evening = volume_evening,
        vol_night = volume_night
    )

    return excel_hours_info

def compute_ph_system(data):
    horas = [x[0] for x in data]
    moda_horas = Counter(horas).most_common()
    if len(moda_horas) == 1:
        result = moda_horas[0][0]
    else:
        max_vol = {}
        for hora, _ in moda_horas:
            vol_per_hour = [volumen for h, volumen in data if h == hora]
            max_vol[hora] = max(vol_per_hour)

        result = max(max_vol, key=max_vol.get)
    return result