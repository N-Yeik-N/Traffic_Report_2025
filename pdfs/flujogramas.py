import os
from pdfs.tools import *
import re
from openpyxl import load_workbook
from docxtpl import DocxTemplate
from docxcompose.composer import Composer

def _combine_all_docx(filePathMaster, filePathsList, finalPath) -> None:
    number_of_sections = len(filePathsList)
    master = Document(filePathMaster)
    composer = Composer(master)
    for i in range(0, number_of_sections):
        doc_temp = Document(filePathsList[i])
        composer.append(doc_temp)
    composer.save(finalPath)

def _read_side_volumes(excelPath: str, maxStage: str)-> dict:    
    wb = load_workbook(excelPath, read_only=True, data_only=True)
    ws = wb[f'V_{maxStage[:2]}']
    volByAccess = dict({(row[0].value, row[1].value) for row in ws["BT35:BU42"]})
    wb.close()

    return volByAccess

def construir_texto_volumenes(volByAccess):
    partes = []

    for acceso, volumen in volByAccess.items():
        if volByAccess.get(acceso, 0) > 0:
            partes.append(f"el acceso {acceso} tiene un volumen vehicular de {volumen} veh/h")
    return ', '.join(partes) + '.' if partes else ''

def flujograma_vehicular(pathSubarea: str, maxStage, maxTipicidad) -> str:
    listCodes = get_codes(pathSubarea)
    anexos_path = os.path.join(pathSubarea, "Anexos")

    folderAnexos = os.listdir(anexos_path)

    assert "Vehicular" in folderAnexos, "ERROR: No se encontro el archivo 'Vehicular' en la carpeta 'Anexos'"

    folderVehicular = os.path.join(anexos_path, "Vehicular")
    listPDFS = os.listdir(folderVehicular)

    pdfs_by_code = {}
    for code in listCodes:
        pdfs_by_code[code] = []

    pattern1 = r"([A-Z]+[0-9]+)"
    pattern2 = r"([A-Z]+-[0-9]+)"
    for pdf in listPDFS:
        match_pdf = re.search(pattern1, pdf) or re.search(pattern2, pdf)
        if match_pdf:
            code_str = match_pdf[1]
            pdfs_by_code[code_str].append(pdf)

    listSelectedPDF = []
    listCodes = []

    findString = "V_"
    if maxStage == "Mañana":
        findString += "Ma_"
    elif maxStage == "Tarde":
        findString += "Ta_"
    elif maxStage == "Noche":
        findString += "No_"

    if maxTipicidad == "típico":
        findString += "T"
    else:
        findString += "A"

    for code, pdfs in pdfs_by_code.items():
        for pdf in pdfs:
            if findString in pdf and not pdf.endswith('.png'):
                listSelectedPDF.append((code, os.path.join(folderVehicular, pdf)))
                listCodes.append(code)

    dataInfo = []
    for code, pdfPath in listSelectedPDF:
        namePDF = os.path.split(pdfPath)[1]
        namePDF = namePDF[:-4]
        dataInfo.append([
            code,
            convert_pdf_to_image(pdfPath, folderVehicular, namePDF),
        ])

    flujogramaPath = create_flujogramas_vehicular_subdocs(dataInfo, pathSubarea, maxStage, maxTipicidad)

    return flujogramaPath

def create_paragraphs(subareaPath: str, maxTipicidad: str, maxStage: str):
    listCodes = get_codes(subareaPath)
    pathParts = subareaPath.split("/")
    proyectFolder = '/'.join(pathParts[:-2])
    subareaName = pathParts[-1]

    fieldPath = os.path.join(
        proyectFolder,
        "7. Informacion de Campo",
        subareaName,
        "Vehicular",
    )

    if maxTipicidad == "típico":
        tipicidad = "Tipico"
    else:
        tipicidad = "Atipico"

    tipicidadFolder = os.path.join(
        fieldPath,
        tipicidad,
    )

    excelFiles = os.listdir(tipicidadFolder)
    excelFiles = [file for file in excelFiles if file.endswith(".xlsm") and not file.startswith("~$")]

    pattern = r"([A-Z]+-[0-9]+)"

    paragraphs = {}
    for code in listCodes:
        paragraphs[code] = {}

    for excelFile in excelFiles:
        match = re.search(pattern, excelFile)
        if match:
            code = match[1]
            if code in listCodes:
                excelPath = os.path.join(tipicidadFolder, excelFile)
                volByAccess = _read_side_volumes(excelPath, maxStage)
                texto_volumenes = construir_texto_volumenes(volByAccess)
                paragraphs[code] = texto_volumenes 

    count = 1
    listParagraphPaths = []
    for code, paragraph in paragraphs.items():
        docTemplate = DocxTemplate("./templates/template_lista.docx")
        docTemplate.render({
            "codinterindividual": code,
            "vol_by_access": paragraph,
        })
        paragraphPath = os.path.join(
            subareaPath,
            "Tablas",
            f"paragraph_flujograma_{count}.docx",
        )
        docTemplate.save(paragraphPath)
        listParagraphPaths.append(paragraphPath)
        count += 1

    finalPath = os.path.join(subareaPath, "Tablas", "flujograma_parrafos.docx")
    if len(listParagraphPaths) > 1:
        filePathMaster = listParagraphPaths[0]
        filePathList = listParagraphPaths[1:]
        _combine_all_docx(filePathMaster, filePathList, finalPath)
    else:
        finalPath = listParagraphPaths[0]

    return finalPath

def flujograma_peatonal(path_subarea, maxStage, maxTipicidad) -> str:
    listCodes = get_codes(path_subarea)
    anexos_path = os.path.join(path_subarea, "Anexos")

    folderAnexos = os.listdir(anexos_path)

    if not "Peatonal" in folderAnexos:
        print("ERROR: No se encontro el archivo 'Vehicular' en la carpeta 'Anexos'")

    folderPeatonal = os.path.join(anexos_path, "Peatonal")
    listPDFS = os.listdir(folderPeatonal)

    pdfs_by_code = {}
    for code in listCodes:
        pdfs_by_code[code] = []

    pattern1 = r"([A-Z]+[0-9]+)"
    pattern2 = r"([A-Z]+-[0-9]+)"
    for pdf in listPDFS:
        match_pdf = re.search(pattern1, pdf) or re.search(pattern2, pdf)
        if match_pdf:
            code_str = match_pdf[1]
            pdfs_by_code[code_str].append(pdf)

    listSelectedPDF = []
    listCodes = []
    for code, pdfs in pdfs_by_code.items():
        for pdf in pdfs:
            if "Turno 01_T" in pdf:
                listSelectedPDF.append((code, os.path.join(folderPeatonal, pdf)))
                listCodes.append(code)
    listCodes = list(set(listCodes))

    listPathImages = {}

    for code in listCodes:
        listPathImages[code] = []

    #Checking if there are images
    for code in listCodes:
        for codePath, pdfPath in listSelectedPDF:
            if code == codePath:
                if pdfPath.endswith('.png'):
                    listPathImages[code].append(pdfPath)
                    break

    #In case there are no .png files
    listPngImages = listPathImages.copy()
    for code, listDocuments in listPngImages.items():
        if len(listDocuments) == 0: #There is no .pngs
            for codePDF, pdfPath in listSelectedPDF:
                if code == codePDF:
                    namePDF = os.path.split(pdfPath)[1]
                    namePDF = namePDF[:-4]
                    listPathImages[code] = convert_pdf_to_image(pdfPath, folderPeatonal, namePDF)
    
    resultList = []
    for code, imagePath in listPathImages.items():
        resultList.append((code, imagePath[0]))

    flujograma_path = create_flujograma_peatonal_subdocs(resultList, path_subarea, maxStage, maxTipicidad)

    return flujograma_path