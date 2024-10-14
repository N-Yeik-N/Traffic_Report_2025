from openpyxl import load_workbook
import matplotlib.pyplot as plt
import csv
import os
from docxtpl import DocxTemplate, InlineImage
from docx.shared import Inches, Cm
from docxcompose.composer import Composer
from docx import Document

def _convert_quarter2hour(hour: str) -> str:
    # Dividir el texto en los límites inferior y superior
    start_time, end_time = hour.split(" - ")

    # Convertir las partes del tiempo en horas y minutos
    start_hour, start_minute = map(int, start_time.split(":"))
    end_hour, end_minute = map(int, end_time.split(":"))

    # Restar una hora del límite superior
    end_hour -= 1

    # Asegurarse de que el nuevo límite superior es una hora válida (0-23)
    if end_hour < 0:
        end_hour += 24

    # Formatear las nuevas horas y minutos en cadenas de texto
    new_start_time = f"{end_hour:02d}:{end_minute:02d}"
    new_end_time = f"{end_hour + 1:02d}:{end_minute:02d}"

    return f"{new_start_time} - {new_end_time}"

def _draw_hist(subareaPath, volumes: list, nameIntersection: str, peakHoursList: list, countImages: int, labels: list, pedestrian: bool = False) -> str:
    # Crear la figura y los ejes
    fig, ax = plt.subplots(figsize=(10, 6))

    volumes = [int(x) for x in volumes]

    # Crear el gráfico de barras
    bars = ax.bar(range(len(volumes)), volumes, color='#A6A6A6')
    selectedIndexes = []
    for i in peakHoursList:
        selectedIndexes.append(
            (i+1, i+2)
            )
        for j in range(i, i+4):
            bars[j].set_color('#1f77b4')

    fullText = "PEATONAL" if pedestrian else "VEHICULAR"
    shortText = "pea" if pedestrian else "veh"
    headers = ["TURNO ", "\nHora Punta Sistema\n"]
    stageDay = ["MAÑANA", "TARDE", "NOCHE"]
    maxTotal = 0
    relevantBars = []
    for (id1, id2), stage in zip(selectedIndexes, stageDay):
        relevantBars.append([
            (id1-1, bars[id1-1].get_height()),
            (id1, bars[id1].get_height()),
            (id2, bars[id2].get_height()),
            (id2+1, bars[id2+1].get_height()),
            ])

    for (id1, id2), stage, relevantBarGroup in zip(selectedIndexes, stageDay, relevantBars):
        midPoint_X = (bars[id1].get_x() + bars[id1].get_width() / 2 + bars[id2].get_x() + bars[id2].get_width() / 2) / 2
        text = headers[0] + stage + '\n' + str(int((tuple(sum(x) for x in zip(*relevantBarGroup))[1]))) + " " + shortText + "s/h" + headers[1] + _convert_quarter2hour(str(labels[relevantBarGroup[-1][0]])) # Copilot sum
        ax.text(midPoint_X, max(volumes)*1.38, text, ha='center', va='center', fontsize = 10, bbox=dict(facecolor='white', alpha=0.5))

    # Girar las etiquetas del eje x
    ax.set_xticks(range(len(labels)))
    ax.set_xticklabels(labels, rotation=90)

    # Obtener el límite máximo actual del eje y
    _, y_max = ax.get_ylim()
    ax.set_ylim(top=y_max*1.5)

    # Añadir los valores encima de las barras
    maxHeight = 0
    for bar in bars:
        maxHeight = max(maxHeight, bar.get_height())

    for bar in bars:
        yval = bar.get_height()
        if yval == 0: continue
        ax.text(bar.get_x() + bar.get_width()/2, 0.04*maxHeight+yval, round(yval, 2), ha='center', va='bottom', color='black', rotation=90)

    # Configurar etiquetas y título
    ax.set_ylabel(f'Volumen ({shortText}/15 min)')
    ax.set_title(f'HISTOGRAMA {fullText}\n{nameIntersection}', fontweight='bold')

    # Mostrar el gráfico
    plt.tight_layout()
    finalPath = os.path.join(subareaPath, "Tablas", f"HistogramaVehicular_{countImages}.png")
    plt.savefig(finalPath)
    plt.close()

    return finalPath

def _get_data(rangeSlice, ws):
    volumes = [row[0].value if row[0].value is not None else 0 for row in ws[rangeSlice]]
    return volumes

def _combine_all_docx(filePathMaster, filePathsList, finalPath) -> None:
    number_of_sections = len(filePathsList)
    master = Document(filePathMaster)
    composer = Composer(master)
    for i in range(0, number_of_sections):
        doc_temp = Document(filePathsList[i])
        composer.append(doc_temp)

    composer.save(finalPath)

def create_histograma_vehicular(
        subareaPath: str,
        excelPath: str, #Ruta del excel a conseguir su histograma
        txtPath: str, #Ruta del .txt con la hora punta del sistema en formato flotante
        countImages: int, #Contador de imágenes de histogramas vehiculares creados
        ) -> None:
    
    #Obtener las horas puntas del sistema desplazados
    peakHoursList = []
    increments = [6.25, 11.75, 17.25]
    with open(txtPath, 'r') as file:
        reader = csv.reader(file, delimiter='\t')
        count = 0
        for row in reader:
            peakHoursList.append(int((float(row[-1])-increments[count])*4))
            count += 1

    wb = load_workbook(excelPath, read_only=True, data_only=True)
    ws = wb['N']

    nameIntersection = wb['Inicio']["G5"].value

    listSlices15minVolumes = [
        slice("HQ41", "HQ54"),
        slice("HQ63", "HQ76"),
        slice("HQ85", "HQ98"),
    ]

    listSlicesLabels = [
        slice("J41", "J54"),
        slice("J63", "J76"),
        slice("J85", "J98"),
    ]

    volumes = []
    labels = []

    separations = []
    for rangeVolumes, rangeLabels in zip(listSlices15minVolumes, listSlicesLabels):
        volume = _get_data(rangeVolumes, ws)
        volumes.extend(volume)
        separations.append(len(volume))

        label = _get_data(rangeLabels, ws)
        labels.extend(label)

    wb.close()

    peakHoursList = [peakHoursList[0],
                    peakHoursList[1] + separations[0],
                    peakHoursList[2] + sum(separations[:2]),
                    ]
    
    finalPath = _draw_hist(
        subareaPath,
        volumes,
        nameIntersection,
        peakHoursList,
        countImages,
        labels,
    )

    return finalPath, nameIntersection, volumes, labels, peakHoursList

def create_histograma_peatonal(
        subareaPath: str,
        excelPath: str, #Ruta del excel a conseguir su histograma
        txtPath: str, #Ruta del .txt con la hora punta del sistema en formato flotante
        countImages: int, #Contador de imágenes de histogramas vehiculares creados
        ) -> None:
    
    #Obtener las horas puntas del sistema desplazados
    peakHoursList = []
    increments = [6.25, 11.75, 17.25]
    with open(txtPath, 'r') as file:
        reader = csv.reader(file, delimiter='\t')
        count = 0
        for row in reader:
            peakHoursList.append(int((float(row[-1])-increments[count])*4))
            count += 1

    wb = load_workbook(excelPath, read_only=True, data_only=True)
    ws = wb['Data Peatonal']

    nameIntersection = wb['Inicio']["G4"].value

    listSlicesVolumes = [
        slice("UZ21", "UZ34"),
        slice("UZ43", "UZ56"),
        slice("UZ65", "UZ78"),
    ]

    listSlicesLabels = [
        slice("K21", "K34"),
        slice("K43", "K56"),
        slice("K65", "K78"),
    ]

    volumes = []
    labels = []

    separations = []
    for rangeVolumes, rangeLabels in zip(listSlicesVolumes, listSlicesLabels):
        volume = _get_data(rangeVolumes, ws)
        volumes.extend(volume)
        separations.append(len(volume))

        label = _get_data(rangeLabels, ws)
        labels.extend(label)

    wb.close()

    peakHoursList = [peakHoursList[0],
                    peakHoursList[1] + separations[0],
                    peakHoursList[2] + sum(separations[:2]),
                    ]
    
    finalPath = _draw_hist(
        subareaPath,
        volumes,
        nameIntersection,
        peakHoursList,
        countImages,
        labels,
        True
    )

    return finalPath, nameIntersection, volumes, labels

def histogramas_vehiculares(subareaPath: str) -> str:
    #List of excels
    pathParts = subareaPath.split("/")
    subareaID = pathParts[-1]
    proyectFolder = '/'.join(pathParts[:-2])
    fieldData = os.path.join(
        proyectFolder,
        "7. Informacion de Campo",
        subareaID,
        "Vehicular",
    )
    
    #Find PeakHours.txt
    txtPaths = {
        "Tipico": None,
        "Atipico": None,
    }
    try:
        txtPaths["Tipico"] = os.path.join(subareaPath, "Tablas", "PeakHoursTipico.txt")
        txtPaths["Atipico"] = os.path.join(subareaPath, "Tablas", "PeakHoursAtipico.txt")
    except FileNotFoundError:
        return print("Error: no existe el archivo PeakHours.txt en la carpeta 'Tablas'")

    wordsByTipicidad = {
        "Tipico": [],
        "Atipico": [],
    }

    systemVolumes = {
        "Tipico": None,
        "Atipico": None,
    }

    totalVolumes ={
        "Tipico": None,
        "Atipico": None,
    }

    countImages = 1
    for tipicidad in ["Tipico", "Atipico"]:
        typicalPath = os.path.join(fieldData, tipicidad)
        excelList = os.listdir(typicalPath)
        excelList = [file for file in excelList if file.endswith(".xlsm") and not file.startswith("~$")]
        VOLUME_SYSTEM = None
        VOLUME_TOTAL = None
        for excel in excelList:
            excelPath = os.path.join(typicalPath, excel)
            histogramaPath, nameIntersection, volumes, labels, peakHoursList = create_histograma_vehicular( #NOTE: labels se sobreescribe una y otra vez.
                subareaPath,
                excelPath,
                txtPaths[tipicidad],
                countImages,
                )
            VOLUME_PARTIAL = [x+y+z for x,y,z in zip(
                                                    volumes[peakHoursList[0]:peakHoursList[0]+4],
                                                    volumes[peakHoursList[1]:peakHoursList[1]+4],
                                                    volumes[peakHoursList[2]:peakHoursList[2]+4],
                                                    )]
            if VOLUME_SYSTEM:
                VOLUME_SYSTEM += VOLUME_PARTIAL
                VOLUME_TOTAL = [x+y for x,y in zip(VOLUME_TOTAL, volumes)]
            else:
                VOLUME_SYSTEM = VOLUME_PARTIAL
                VOLUME_TOTAL = volumes
            
            if tipicidad == "Tipico":
                tipicidadTxt = "típico"
            else:
                tipicidadTxt = "atípico"
            texto = f"Histograma vehicular {tipicidadTxt} de la {nameIntersection}"

            docTemplate = DocxTemplate("./templates/template_imagenes.docx")
            newImage = InlineImage(docTemplate, histogramaPath, height=Cm(7.5))
            docTemplate.render({"texto": texto, "tabla": newImage})

            finalPath = os.path.join(subareaPath, "Tablas", f"HistogramaVehicular_{tipicidad}_{countImages}.docx")
            docTemplate.save(finalPath)
            wordsByTipicidad[tipicidad].append(finalPath)
            countImages += 1
        totalVolumes[tipicidad] = VOLUME_TOTAL
        systemVolumes[tipicidad] = VOLUME_SYSTEM

    #Volúmenes totales
    volumesByStages = {
        "Tipico": {
            "Mañana": None,
            "Tarde": None,
            "Noche": None,
        },
        "Atipico": {
            "Mañana": None,
            "Tarde": None,
            "Noche": None,
        }
    }

    for tipicidad, VOLUME_TOTAL in totalVolumes.items():
        volumesByStages[tipicidad]["Mañana"] = sum(VOLUME_TOTAL[0:14])
        volumesByStages[tipicidad]["Tarde"]  = sum(VOLUME_TOTAL[14:28])
        volumesByStages[tipicidad]["Noche"] = sum(VOLUME_TOTAL[28:])

    for tipicidad, VOLUME in systemVolumes.items():
        if tipicidad == "Tipico":
            sumvoltip = sum(VOLUME)
        else:
            sumvolati = sum(VOLUME)

    maxStageByTipicidad = {}
    for tipicidad, stages in volumesByStages.items():
        maxStage = max(stages, key=stages.get)
        maxStageByTipicidad[tipicidad] = maxStage
        
    if sumvoltip > sumvolati:
        maxtipicidad = "típico"
        volturnmanana = str(volumesByStages["Tipico"]["Mañana"])
        volturntarde = str(volumesByStages["Tipico"]["Tarde"])
        volturnnoche = str(volumesByStages["Tipico"]["Noche"])
        maxturno = str(maxStageByTipicidad["Tipico"])
    else:
        maxtipicidad = "atípico"
        volturnmanana = str(volumesByStages["Atipico"]["Mañana"])
        volturntarde = str(volumesByStages["Atipico"]["Tarde"])
        volturnnoche = str(volumesByStages["Atipico"]["Noche"])
        maxturno = str(maxStageByTipicidad["Atipico"])

    #peakHourList for totals
    pathTotalHist = {
        "Tipico": None,
        "Atipico": None,
    }

    separations = [14,14,14]
    increments = [6.25, 11.75, 17.25]
    for tipicidad, VOLUME in totalVolumes.items():
        peakHoursList = []
        with open(txtPaths[tipicidad], 'r') as file:
            reader = csv.reader(file, delimiter='\t')
            count = 0
            for row in reader:
                peakHoursList.append(int((float(row[-1])-increments[count])*4))
                count += 1

        peakHoursList = [
            peakHoursList[0],
            peakHoursList[1] + separations[0],
            peakHoursList[2] + sum(separations[:2])
        ]

        pathTotal = _draw_hist(
            subareaPath,
            VOLUME,
            "SISTEMA",
            peakHoursList,
            countImages,
            labels,
        )

        if tipicidad == "Tipico":
            tipicidadTxt = "típico"
        else:
            tipicidadTxt = "atípico"
        docTemplate = DocxTemplate("./templates/template_imagenes.docx")
        newImage = InlineImage(docTemplate, pathTotal, width=Inches(6))
        docTemplate.render({"texto": f"Histograma vehicular {tipicidadTxt} del sistema", "tabla": newImage})

        finalPath = os.path.join(subareaPath, "Tablas", f"HistogramaVehicular_{tipicidad}_SISTEMA.docx")
        docTemplate.save(finalPath)
        pathTotalHist[tipicidad] = finalPath

        countImages += 1

    histogramaPathByTipicidad = {
        "Tipico": None,
        "Atipico": None,
    }

    for key, listPaths in wordsByTipicidad.items():
        filePathMaster = listPaths[0]
        filePathList = listPaths[1:]
        histogramaDocx = os.path.join(subareaPath, "Tablas", f"HistogramaVehicular_{key}.docx")
        _combine_all_docx(filePathMaster, filePathList, histogramaDocx)
        histogramaPathByTipicidad[key] = histogramaDocx

    return histogramaPathByTipicidad["Tipico"], histogramaPathByTipicidad["Atipico"], pathTotalHist["Tipico"], pathTotalHist["Atipico"], sumvoltip, sumvolati, maxtipicidad, volturnmanana, volturntarde, volturnnoche, maxturno

def histogramas_peatonales(subareaPath: str) -> str:
    #List of excels
    pathParts = subareaPath.split("/")
    subareaID = pathParts[-1]
    proyectFolder = '/'.join(pathParts[:-2])
    fieldData = os.path.join(
        proyectFolder,
        "7. Informacion de Campo",
        subareaID,
        "Peatonal",
    )

    #Find PeakHours.txt
    txtPaths = {
        "Tipico": None,
        "Atipico": None,
    }
    try:
        txtPaths["Tipico"] = os.path.join(subareaPath, "Tablas", "PeakHoursTipico.txt")
        txtPaths["Atipico"] = os.path.join(subareaPath, "Tablas", "PeakHoursAtipico.txt")
    except FileNotFoundError:
        return print("Error: no existe el archivo PeakHours.txt en la carpeta 'Tablas'")

    wordsByTipicidad = {
        "Tipico": [],
        "Atipico": [],
    }

    countImages = 1
    for tipicidad in ["Tipico", "Atipico"]:
        typicalPath = os.path.join(fieldData, tipicidad)
        excelList = os.listdir(typicalPath)
        excelList = [file for file in excelList if file.endswith(".xlsm") and not file.startswith("~$")]
        for excel in excelList:
            excelPath = os.path.join(typicalPath, excel)
            histogramaPath, nameIntersection, volumes, labels = create_histograma_peatonal( #NOTE: labels se sobreescribe una y otra vez.
                subareaPath,
                excelPath,
                txtPaths[tipicidad],
                countImages,
                )
            
            if tipicidad == "Tipico":
                tipicidadTxt = "típico"
            else:
                tipicidadTxt = "atípico"

            texto = f"Histograma peatonal {tipicidadTxt} de la {nameIntersection}"

            docTemplate = DocxTemplate("./templates/template_imagenes.docx")
            newImage = InlineImage(docTemplate, histogramaPath, height=Cm(7.5))
            docTemplate.render({"texto": texto, "tabla": newImage})

            finalPath = os.path.join(subareaPath, "Tablas", f"HistogramaVehicular_{tipicidad}_{countImages}.docx")
            docTemplate.save(finalPath)
            wordsByTipicidad[tipicidad].append(finalPath)
            countImages += 1

    histogramaPathByTipicidad = {
        "Tipico": None,
        "Atipico": None,
    }

    for key, listPaths in wordsByTipicidad.items():
        if len(listPaths) > 1:
            filePathMaster = listPaths[0]
            filePathList = listPaths[1:]
            histogramaDocx = os.path.join(subareaPath, "Tablas", f"Histograma_Peatonal_{key}.docx")
            _combine_all_docx(filePathMaster, filePathList, histogramaDocx)
        else:
            histogramaDocx = listPaths[0]
        histogramaPathByTipicidad[key] = histogramaDocx

    return histogramaPathByTipicidad["Tipico"], histogramaPathByTipicidad["Atipico"]