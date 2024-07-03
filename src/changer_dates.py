import win32com.client as com
from openpyxl import load_workbook
import math
import os

def _change_dates(sheet: object, celdas: list, valores: list) -> None:
    for celda, valor in zip(celdas, valores):
        sheet.Range(celda).Value = valor

def change_peakhours(excel, excelPath: str) -> None:
    #Lectura de datos
    slicesTurns = [
        slice("HR40", "HR63"), #06:00 - 12:00
        slice("HR64", "HR83"), #12:00 - 17:00
        slice("HR84", "HR111"), #17:00 - 00:00
    ]

    turns = ["Mañana", "Tarde", "Noche"]

    locations = [
        ["C18", "G18"],
        ["C25", "G25"],
        ["C32", "G32"],
    ]

    indexExcel = [25, 49, 69]

    wb = load_workbook(excelPath, read_only=True, data_only=True)
    ws = wb['N']
    volumesByTurn = {}
    for sliceTurn, turn in zip(slicesTurns, turns):
        data = [row[0].value for row in ws[sliceTurn]]
        data = [0 if isinstance(x, str) or math.isnan(x) else x for x in data]
        volumesByTurn[turn] = data

    wb.close()

    #Ingreso de datos
    workbook = excel.Workbooks.Open(excelPath)
    sheet = workbook.Worksheets("Histograma")

    for (turn, listVolumes), index, celdas in zip(volumesByTurn.items(), indexExcel, locations):
        maxValue = max(listVolumes)
        indexMax = listVolumes.index(maxValue)
        hora = (indexMax+index)//4
        minuto = int(((indexMax+index)%4)*15)
        upperHour = "{:02d}:{:02d}".format(hora, minuto)
        lowerHour = "{:02d}:{:02d}".format(hora-1, minuto)

        try:
            _change_dates(
                sheet,
                celdas,
                [lowerHour, upperHour],
            )
        except Exception as inst:
            print("Error: ", inst)
            print("Turno: ", turn)
            continue

    workbook.Save()
    workbook.Close(SaveChanges = True)