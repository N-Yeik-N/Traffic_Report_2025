import matplotlib.pyplot as plt
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
import os
from unidecode import unidecode

def _diana(
        datos: list,
        texto2: str,
        texto3: str,
        texto4: str,
        path: str | Path,
) -> None:
    # Definir los límites de los anillos
    anillos = [(0, 5, 'darkgreen'), (5, 10, 'lightgreen'), (10, 15, 'yellow'), (15, 20, 'lightsalmon')]

    # Crear la figura y los ejes
    fig, ax = plt.subplots()

    # Dibujar los anillos
    for i, (lim_inf, lim_sup, color) in enumerate(anillos):
        ax.add_patch(plt.Circle((0, 0), lim_sup, color=color, alpha=0.2))
        ax.annotate(f'', xy=(0, lim_sup-0.5), ha='center')

    # Convertir los resultados GEH en radios
    radios = datos

    # Generar ángulos aleatorios en coordenadas polares para los datos de GEH
    angulos = np.random.uniform(0, 2 * np.pi, size=len(datos))

    # Convertir coordenadas polares a cartesianas
    x = radios * np.cos(angulos)
    y = radios * np.sin(angulos)

    # Dibujar los datos como puntos en la diana con rotaciones aleatorias
    ax.scatter(x, y, color='blue', alpha=0.5, marker='x')

    # Establecer los límites y valores de las marcas de los ejes
    limite_max = 20
    limite_min = -20
    intervalo = 5
    ax.set_xlim(limite_min, limite_max)
    ax.set_ylim(limite_min, limite_max)
    ax.set_xticks(np.arange(limite_min, limite_max + intervalo, intervalo))
    ax.set_yticks(np.arange(limite_min, limite_max + intervalo, intervalo))

    # Agregar línea en X = 0 y etiquetas cada 5 en 5
    ax.axhline(0, color='gray', linewidth=0.5)
    for i in range(-5+limite_min + intervalo, limite_max+5, intervalo):
        ax.text(0, i, str(abs(i)), ha='right', va='center', fontsize=8)

    # Agregar línea en Y = 0
    ax.axvline(0, color='gray', linewidth=0.5)

    # Agregar grilla de color gris
    ax.grid(color='grey', linestyle='--', linewidth=0.5)

    # Agregar texto en la esquina superior izquierda
    texto1 = 'Diana GEH'
    ax.text(limite_min + 1, limite_max - 2, texto1, fontsize=12, color='black', fontweight='bold')
    ax.text(limite_min + 1, limite_max - 4, texto2, fontsize=10, color='black', fontweight='bold')
    ax.text(limite_min + 1, limite_max - 6, texto3, fontsize=10, color='black', fontweight='bold')
    ax.text(limite_min + 1, limite_max - 8, texto4, fontsize=10, color='black', fontweight='bold')

    # Eliminar los ejes
    ax.xaxis.set_label_position('bottom')
    ax.yaxis.set_label_position('left')
    ax.spines['left'].set_position(('outward', 0))
    ax.spines['bottom'].set_position(('outward', 0))

    # Cambiar el color de los ticks a gris
    ax.tick_params(axis='both', colors='grey')

    for spine in ax.spines.values():
        spine.set_color('grey')

    # Mostrar el gráfico
    plt.gca().set_aspect('equal', adjustable='box')
    plt.title('GEH')
    if not os.path.exists(path):
        os.makedirs(path)
    finalName = f'GEH_{unidecode(texto2[10:]).upper()}.png'
    plt.savefig(os.path.join(path, finalName), dpi=300, bbox_inches = 'tight')
    plt.close()

def create_dianas(excelPath) -> None:
    wb = load_workbook(excelPath, read_only=True, data_only=True)
    ws = wb['GEH']
    valores = [[elem.value for elem in row] for row in ws[slice("BW7", "CP7")]][0]
    num_veh = next((i for i, val in enumerate(valores) if val in ('n', 'N', 1, None)), None)

    num_giros = [row[0].value for row in ws[slice("K8","K500")]].index(None)+1
    list_data = []
    for j in range(num_veh):
        data = []
        for i in range(num_giros):
            valor = ws.cell(row = 7+i, column = 75 +j).value
            if type(valor) == str:
                data.append(valor)
            elif type(valor) == float or int:
                data.append(round(valor, 1))
        list_data.append(data)

    wb.close()

    pathParts = excelPath.split('\\')
    subareaPath = pathParts[:-4]
    subareaPath = '/'.join(subareaPath)
    directory = os.path.join(subareaPath, 'Tablas')

    for index, data in enumerate(list_data):
        less_five = 0
        for elem in data[1:]:
            if elem < 5:
                less_five += 1
        percentage = (less_five/len(data[1:]))*100
        _diana(
            datos = data[1:],
            texto2 = f'Veh Type: {data[0]}',
            texto3 = f'Muestra: {int((num_giros))}',
            texto4 = f'GEH<5: {percentage:.1f}%',
            path = directory,
        )

    wb.close()