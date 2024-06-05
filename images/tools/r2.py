import matplotlib.pyplot as plt
from openpyxl import load_workbook
import numpy as np
from scipy import stats
import pandas as pd
import os
from unidecode import unidecode

def _r2_figure(CAMPO, MODELO, vehicularType, PATH) -> None:
    # Convertir listas a arrays de numpy
    CAMPO = np.array(CAMPO)
    MODELO = np.array(MODELO)
    
    # Verificar si todos los valores son cero
    if np.all(CAMPO == 0) and np.all(MODELO == 0):
        plt.scatter(CAMPO, MODELO, label=f"Tipo {vehicularType}")
        plt.xlabel('Volumen campo', fontweight='bold', fontsize = '14')
        plt.ylabel('Volumen modelo', fontweight='bold', fontsize = '14')
        plt.title(f'R2: Volumen "{vehicularType}" (Datos en cero)', fontweight='bold', fontsize = '16')
        plt.legend()
        plt.grid(True)
        plt.savefig(PATH)
        plt.close()
        return

    try:
        slope, intercept, r_value, p_value, std_err = stats.linregress(CAMPO, MODELO)
    except ValueError as e:
        return None
    except Exception as e:
        raise e

    fieldRange = np.linspace(min(CAMPO), max(CAMPO), 100)
    model_pred = slope * fieldRange + intercept

    plt.scatter(CAMPO, MODELO, label=f"Tipo {vehicularType}")
    plt.plot(fieldRange, model_pred, color='blue', label='Regresión Lineal')
    plt.xlabel('Volumen campo', fontweight='bold', fontsize = '14')
    plt.ylabel('Volumen modelo', fontweight='bold', fontsize = '14')
    plt.title(f'R2: Volumen "{vehicularType}"', fontweight='bold', fontsize = '16')
    r_value_str = str(r_value)
    try:
        entero, decimal = r_value_str.split('.')
        r_value_show = f'{entero}.{decimal[:2]}'
    except ValueError as inst:
        r_value_show = r_value_str
    plt.text(0.8, 0.1, f'$R^2 = {r_value_show}$\ny={slope:.2f}x+({intercept:.2f})', ha='center', va='center', transform=plt.gca().transAxes, bbox=dict(facecolor="white", alpha=0.5))

    plt.legend()
    plt.grid(True)
    plt.savefig(PATH)
    plt.close()

def create_r2s(excelPath) -> None:
    wb = load_workbook(excelPath, read_only=True, data_only=True)
    ws = wb.active

    typesNumber = next((i for i, val in enumerate(row[0].value for row in ws[slice('H8', 'H38')]) if val in ('n', 'N', None)), None)
    typesNames = [row[0].value for row in ws[slice('H8','H38')]][:typesNumber]
    nameCols = [row[0].value for row in ws[slice('H8', 'H38')]][:typesNumber]
    odNumber = [row[0].value for row in ws[slice('K8','K1009')]].index(None)
    wb.close()

    dfCampo = pd.read_excel(excelPath, sheet_name='GEH', usecols='AI:BB', nrows=odNumber, skiprows=6)
    dfCampo = dfCampo.iloc[:, :typesNumber]
    dfCampo.columns = nameCols

    dfModelo = pd.read_excel(excelPath, sheet_name='GEH', usecols='BC:BV', nrows=odNumber, skiprows=6)
    dfModelo = dfModelo.iloc[:, :typesNumber]
    dfModelo.columns = nameCols

    pathParts = excelPath.split('\\')
    subareaPath = pathParts[:-4]
    subareaPath = '/'.join(subareaPath)

    for i in range(len(dfCampo.columns)):
        CAMPO = list(dfCampo.iloc[:,i])
        MODELO = list(dfModelo.iloc[:,i])
        typeName = typesNames[i]
        finalName = f'R2_{unidecode(typeName).upper()}.png'
        finalPath = os.path.join(subareaPath, 'Tablas', finalName)
        _r2_figure(CAMPO, MODELO, typeName, finalPath)

    wb.close()