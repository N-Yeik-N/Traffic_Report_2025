import pandas as pd
from openpyxl import load_workbook

def board_by_excel(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb['Base Data']
    codigo = ws['C4'].value
    fecha = ws['C5'].value
    columnas = ["Turno", "Sentido", "Acceso", "Tipo de Vehiculo", "Tiempo"]
    df = pd.DataFrame([[cell.value for cell in row] for row in ws['C8:G157']], columns= columnas)
    wb.close()

    dict_info = {
        "Código": codigo, #str
        "Fecha": fecha, #datetime
        "Datos": df, #pd.DataFrame
    }

    return dict_info

if __name__ == '__main__':
    PATH = r"C:\Users\dacan\OneDrive\Desktop\PRUEBAS\Maxima Entropia\04 Proyecto Universitaria (37 Int. - 19 SA)\7. Informacion de Campo\Sub Area 016\Embarque y Desembarque\Tipico\SS-77_Av. Rosa Lozano - Jr. Geranios_T.xlsx"
    dict_info = board_by_excel(PATH)
    print(dict_info)