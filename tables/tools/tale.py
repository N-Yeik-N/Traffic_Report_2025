from openpyxl import load_workbook
import pandas as pd
import re

def tale_by_excel(path):
    pattern1 = r"([A-Z]+[0-9]+)"
    pattern2 = r"([A-Z]+-[0-9]+)"
    name_excel = path.split("\\")[-1]
    coincidence1 = re.search(pattern1, name_excel)
    coincidence2 = re.search(pattern2, name_excel)

    if coincidence1:
        code = coincidence1.group(1)
    elif coincidence2:
        code = coincidence2.group(1)
    else:
        print(f"Error - Este excel no tiene código: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb['Base Data']
    sentido_slice = slice("C7","C14")
    acceso_slice = slice("D7","D14")

    list_sentido = [row[0].value for row in ws[sentido_slice] if row[0].value != None]
    list_acceso = [row[0].value for row in ws[acceso_slice] if row[0].value != None]

    if len(list_acceso) != len(list_sentido): print(f"Error - No coinciden cantidades de sentido-acceso:\n{path}")

    sentido_acceso_dict = {}
    for sent, access in zip(list_sentido, list_acceso):
        sentido_acceso_dict[sent] = access

    quantity = len(list_acceso)

    list_slices = [
        slice("D20","K59"),
        slice("O20","V59"),
        slice("AK20","AR59"),
    ]

    column_names = ["Turn","Direction","Access","Max","Mean","Std"]
    df = pd.DataFrame(columns=column_names)

    for id, turno in enumerate(["Mañana","Tarde","Noche"]):
        means,maxs,stds = _get_statistics(list_slices[id], ws, quantity)
        for acceso, sentido, mean, max, std in zip(list_acceso, list_sentido, means, maxs, stds):
            df.loc[len(df)] = [turno, sentido, acceso, max, mean, std]
            #df = df.concat({"Turn": turno, "Direction": sentido, "Access": acceso, "Max": max, "Mean": mean, "Std": std}, ignore_index=True)

    date = ws['D4'].value

    wb.close()

    return code, date, df

def _get_statistics(slice_turn, ws, quant):
    df = pd.DataFrame([[cell.value for cell in row] for row in ws[slice_turn]]).iloc[:,:quant]
    means = df.mean(axis=0, skipna=True).round(2).fillna('-')
    maxs = df.max(axis=0, skipna=True).round(2).fillna('-')
    stds = df.std(axis=0, skipna=True).round(2).fillna('-')
    return means, maxs,stds