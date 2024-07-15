from dataclasses import dataclass
from openpyxl import load_workbook
import pandas as pd
import os

@dataclass
class PedestrianInfo:
    code: str
    name: str
    idMorning: int
    idEvening: int
    idNight: int
    morningVolume: int
    eveningVolume: int
    nightVolume: int

@dataclass
class PedestrianVolumes:
    codigo: str
    volTotal: int

def str2hour(id: float):
    hour = id//4
    minutes = id%4*15
    formatedHour = "{:02d}:{:02d} - {:02d}:{:02d}".format(hour, minutes, hour+1, minutes)

    return formatedHour

def read_ped_excel(excelPath):
    wb = load_workbook(excelPath, read_only=True, data_only=True)
    ws = wb['Data Peatonal']

    hourSums = [row[0].value for row in ws[slice("VA20","VA83")]]
    hourSums = [x if isinstance(x, int) else 0 for x in hourSums]

    for _ in range(6):
        hourSums[:0] = [0,0,0,0]

    for _ in range(2):
        hourSums.extend([0,0,0,0])
    wb.close()

    return hourSums

def code_by_subarea(path_subarea):
    path_excel = r"data\Cronograma Vissim.xlsx"
    df = pd.read_excel(path_excel, sheet_name='Cronograma', header=0, usecols="A:E", skiprows=1)
    df = df.drop(columns=["Codigo TDR", "Entregable"])

    numsubarea = os.path.split(path_subarea)[1][-3:]
    no = int(numsubarea)

    code_by_subarea = df[df['Sub Area'] == no]['Codigo'].tolist()

    return code_by_subarea