from openpyxl import load_workbook
import re
from dataclasses import dataclass

@dataclass
class cycleTime:
    codigo: str
    nombre: str
    cycleTimeData: list
    phasesData: list

def _get_phases(worksheet: object): #i: 0 = mañana, 1 = tarde, 2 = noche
    cycleTimes = []
    phasesList = []
    for i in [10,11,12,18,19,20]:
        cycleTimes.append(int(worksheet[f'E{i}'].value))
        numPhases = [[elem.value for elem in row] for row in worksheet[slice(f"H{i}", f"AK{i}")]][0].index(None)//3
        phases = [[elem.value for elem in row] for row in worksheet[slice(f"H{i}", f"AK{i}")]][0][:numPhases*3]
        phasesList.append([phases[i:i+3] for i in range(0, len(phases), 3)])

    return cycleTimes, phasesList

def get_info(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    codigo = ws['D3'].value
    pattern1 = r'([A-Z]+-[0-9]+)'
    pattern2 = r'([A-Z]+[0-9]+)'
    coincidence1 = re.search(pattern1, codigo)
    coincidence2 = re.search(pattern2, codigo)
    if coincidence1:
        codinterseccion = coincidence1.group(1)
    elif coincidence2:
        codinterseccion = coincidence2.group(1)
    else:
        print("ERROR - Excel sin código de tipo AA-99 o AA99:\n",path)
    
    nominterseccion = ws['D4'].value
    parts = nominterseccion.split(":")
    nominterseccion = parts[1].strip()

    cycleTimes, phasesList = _get_phases(ws)

    wb.close()

    data = cycleTime(
        codigo = codinterseccion,
        nombre = nominterseccion,
        cycleTimeData = cycleTimes,
        phasesData = phasesList,
    )

    return data