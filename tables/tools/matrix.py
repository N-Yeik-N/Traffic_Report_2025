from openpyxl import load_workbook
import pandas as pd

def read_matrix(excel_path):
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    
    sliceDestiny = slice("B1","AZ1") #Maximum 50 origins
    sliceOrigin = slice("A2", "A51") #Maximum 50 detinys

    origins = ['O'+str(elem.value) for row in ws[sliceOrigin] for elem in row if elem.value != None]
    destinys = ['D'+str(row[0].value) for row in ws[sliceDestiny] for elem in row if elem.value != None]

    numCols = len(destinys)
    numRows = len(origins)

    MATRIX = []
    for row in range(2,numRows+2):
        ROW = []
        for col in range(2,numCols+2):
            cell_value = str(ws.cell(row=row, column=col).value) if ws.cell(row=row, column=col).value != None else ""
            ROW.append(cell_value)
        MATRIX.append(ROW)
    wb.close()

    return origins, destinys, MATRIX