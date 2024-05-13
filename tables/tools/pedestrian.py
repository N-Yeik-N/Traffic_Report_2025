from openpyxl import load_workbook

def get_dates(excel_path):
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb['Inicio']

    date = ws['G5'].value
    wb.close()
    date = date.strftime("%d/%m/%Y")
    
    return date